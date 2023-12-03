import os
import shutil
from typing import Dict, Text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
import PIL.Image
from typing import Text
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from src.draw_text import DrawText
from src.config import (
    PADDING,
    FD_DRAW_TEXT,
    FD_ROTATE,
    LOG,
    FONT_TEXT,
    MAX_HEIGH
)
from datetime import datetime
from src.cli import get_cli
from tqdm import tqdm


class InsertImageToExcel(object):
    def __init__(
        self,
        path_img: Text,
        path_text: Text,
        path_excel: Text
    )->None:
        self.path_img = path_img
        self.path_text = path_text
        self.path_excel = path_excel

        self.lst_img = os.listdir(path = self.path_img)
        self.data :Dict = self.read_data()

        # region create workspace for excel
        self.wb = Workbook()
        self.ws = self.wb.active
        # endregion


    def get_max_id(self):
        # return len(self.ws[f'A{PADDING}'].width)
        return 20

    def get_max_fn(self):
        # return len(self.ws[f'B{PADDING}'].width)
        return 50
    
    def get_max_txt(self):
        # return len(self.ws[f'E{PADDING}'].width)
        return 150
    
    def get_max_img_width(self):
        return 30
        
    def read_data(self):
        with open(self.path_text, 'r', encoding='utf-8') as f:
            data = f.readlines()
        
        data = dict(list(map(lambda x: x.strip().split('\t'), data)))

        return data

    @staticmethod
    def scale_img(img: PIL.Image)->PIL.Image:
        h, w = img.height, img.width

        ratio_h = MAX_HEIGH/ h

        new_h = int(h*ratio_h)
        new_w = int(w*ratio_h)

        new_img = img.resize(size = (new_w, new_h))

        return new_img, new_h, new_w


    
    def rotate_image(
        self,
        path: Text
    )->Image:
        _img = PIL.Image.open(path).rotate(-90, expand=True)

        # region scale image
        scaled_img, h, w = self.scale_img(
            img = _img
        )
        # endregion

        name = os.path.basename(path)
        rotate_path  = os.path.join(FD_ROTATE, name)

        scaled_img.save(
            fp = rotate_path
        )
        return Image(rotate_path), h, w

    def write_img_text(
        self,
        img_name: Text,
        text: Text,
        idx: int
    )->None:
        cell_location_id = f'A{idx+PADDING}'
        cell_location_fn = f'B{idx+PADDING}'
        cell_location_img = f'C{idx+PADDING}'
        cell_location_draw_text = f'D{idx+PADDING}'
        cell_location_text = f'E{idx+PADDING}'

        self.ws[cell_location_id] = str(idx).zfill(4)
        self.ws[cell_location_fn] = img_name

        path = os.path.join(
            self.path_img,
            img_name
        )

        img_rotate, h, w = self.rotate_image(path)

        img_draw_text = DrawText(text, path, h, w)()

        self.ws.add_image(img_rotate , cell_location_img)
        self.ws.add_image(img_draw_text, cell_location_draw_text )
        self.ws[cell_location_text] = text
        
        # Điều chỉnh chiều cao của dòng để phù hợp với kích thước ảnh
        self.ws.row_dimensions[idx+PADDING].height = h
    
    def change_size_text(self):
        for row in tqdm(self.ws.iter_rows(min_row= PADDING, min_col=1, max_col = 2), desc= "Process to Format the cells ID, Filename: "):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=48)
            
        for row in tqdm(self.ws.iter_rows(min_row= PADDING, min_col=5, max_col = 5), desc= "Process to Format the cells Text: "):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=48)

        for row in tqdm(self.ws.iter_rows(min_row= PADDING, min_col=1, max_col = 5), desc= "Process to Format the border: "):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def adjust_col(self):
        self.ws.column_dimensions['A'].width = self.get_max_id() + 4
        self.ws.column_dimensions['B'].width = self.get_max_fn() + 10
        self.ws.column_dimensions['C'].width = self.get_max_img_width() + 10
        self.ws.column_dimensions['D'].width = self.get_max_img_width() + 10
        self.ws.column_dimensions['E'].width = self.get_max_txt() + 10

    def deco_title(self):
        for row in tqdm(self.ws.iter_rows(min_row= 1, min_col=1, max_col = 5, max_row = 1), desc= "Process to Format the cells Header: "):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=35)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.fill = PatternFill(start_color= "008080", end_color="008080", fill_type="solid")

    def process(self):
        # region Get title
        self.ws['A1'] = 'ID'
        self.ws['B1'] = 'FN'
        self.ws['C1'] = 'SRC_IMG'
        self.ws['D1'] = 'TEXT_PRED'
        self.ws['E1'] = 'TEXT'

        # endregion
        for idx, (img_name, text) in tqdm(enumerate(self.data.items()), desc = 'Progress: '):
            try:
                self.write_img_text(
                    img_name,
                    text, 
                    idx
                )
            except Exception as e:
                day_time = datetime.now()
                curr_time = day_time.strftime('%Y%m%d')
                with open(f'{LOG}/{curr_time}.log', 'a', encoding='utf-8') as f:
                    f.write(f'{img_name}\t{text}\t{e}\n')
                continue


    def __call__(self):
        self.process()
        self.deco_title()
        self.change_size_text()
        self.adjust_col()
        self.wb.save(self.path_excel)




if __name__ == '__main__':
    args = get_cli()

    PATH_TEXT = args['path_text']
    PATH_IMG = args['path_img']
    PATH_EXCEL = args['path_excel']

    obj = InsertImageToExcel(
        path_img= PATH_IMG,
        path_text= PATH_TEXT,
        path_excel= PATH_EXCEL
    )

    obj()